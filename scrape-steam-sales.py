import bs4 as bs
import urllib.request
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os.path
from datetime import date

saleData = []
pageCount = 0

# Function for scraping an amazon product. Prints price and a message related to the rating of the object.
def scrapeSteamDailySalesPage(page):

    url='https://store.steampowered.com/search/?specials=1&page='+str(page)
    # user-agent is necessary to get past amazon's non-browser traffic block.
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'
    request = urllib.request.Request(url,headers={'User-Agent': user_agent})

    try:
        source = urllib.request.urlopen(request).read()
    except :
        print(url)
        return
    soup = bs.BeautifulSoup(source,'lxml')

    #add todays date as column
    games = soup.select('.search_result_row')
    if (len(games) > 0) and page < 250 :
        for game in games:
            gameName = game.select('span[class="title"]')[0].text
            gameUrl = game.get('href')
            gameId = gameUrl.split('/',10)[4]
            try:
                gameFullPrice = float(game.select('.search_price')[0].text.split('$',2)[1].strip())
                gameSalePrice = float(game.select('.search_price')[0].text.split('$',2)[2].strip())
                gamePercentOff = (gameFullPrice-gameSalePrice)*100/gameFullPrice
            except:
                continue
            saleData.append([gameName,gameUrl,gameId,gameFullPrice,gameSalePrice,gamePercentOff])



        scrapeSteamDailySalesPage(page+1)

    else :
        print("Scrape complete. Page count: "+str(page))
        updateSpreadsheet()
        return;

def updateSpreadsheet():

    ##Sub functions
    def findFirstEmptyRow():
        for cell in sheet["A"]:
            if cell.value is None :
                return cell.row
        return len(sheet["A"])+1
    def findIdRow(gameId):

        for cell in sheet["C"]:
            if cell.value == gameId:
                return cell.row
        return findFirstEmptyRow()

    filename = "steam-sale-data.xlsx"

    #initalize spreadsheet
    if os.path.exists(filename) :
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
    else :
        workbook = Workbook()
        sheet = workbook.active

        # Setting the headers of the data sheet
        sheet["A1"] = "Name"
        sheet["B1"] = "URL"
        sheet["C1"] = "ID"
        sheet["D1"] = "Price"
        # sheet.column_dimensions["D"].number_format = '$0.00'

    columnToday=len(sheet[1])+1
    # sheet.column_dimensions[get_column_letter(columnToday)].number_format = '$0.00'
    # sheet.cell(row = 1, column= columnToday).number_format = '%Y-%m-%d'
    sheet.cell(row = 1, column= columnToday).value = date.today()

    for game in saleData:
        gameRow = findIdRow(gameId=game[2])
        sheet.cell(row=gameRow, column=1).value = game[0]
        sheet.cell(row=gameRow, column=2).value = game[1]
        sheet.cell(row=gameRow, column=3).value = game[2]
        sheet.cell(row=gameRow, column=4).value = float(game[3])
        sheet.cell(row=gameRow, column=columnToday).value = float(game[4])

    # sheet["D"].number_format = '$0.00''

    workbook.save(filename=filename)

scrapeSteamDailySalesPage(page=1)
